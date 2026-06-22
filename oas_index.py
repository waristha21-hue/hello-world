# -*- coding: utf-8 -*-
"""สร้างดัชนีค้นหา (Excel) ของไฟล์ทั้งหมดใน E:\\diskdisk
คอลัมน์: Section | Brand | Subfolder | Filename | Type | Size(KB) | Full Path
มี AutoFilter + freeze header ให้กรอง/ค้นหาได้ใน Excel
"""
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
ROOT = Path(r"E:\diskdisk")
OUT = ROOT / "_INDEX_OAS.xlsx"

SECTION_NAME = {
    "_Contract": "Contract",
    "_Condition_BKD": "Condition BKD",
}


def classify(rel_parts):
    """คืน (section, brand, subfolder)"""
    top = rel_parts[0]
    if top in SECTION_NAME:
        section = SECTION_NAME[top]
        brand = rel_parts[1] if len(rel_parts) > 2 else "(root)"
        subfolder = "/".join(rel_parts[1:-1])
    else:
        section = "Sale Marketing"
        brand = top
        subfolder = "/".join(rel_parts[:-1])
    return section, brand, subfolder


def main():
    rows = []
    for p in ROOT.rglob("*"):
        if not p.is_file():
            continue
        if p.name.startswith("_INDEX_OAS") or p.suffix.lower() == ".part":
            continue
        rel = p.relative_to(ROOT)
        parts = list(rel.parts)
        section, brand, subfolder = classify(parts)
        size_kb = round(p.stat().st_size / 1024, 1)
        rows.append([section, brand, subfolder, p.name,
                     p.suffix.lower().lstrip(".") or "(none)", size_kb, str(p)])

    rows.sort(key=lambda r: (r[0], r[1], r[2], r[3]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OAS Index"
    headers = ["Section", "Brand", "Subfolder", "Filename", "Type", "Size (KB)", "Full Path"]
    ws.append(headers)

    hfill = PatternFill("solid", fgColor="1F4E78")
    hfont = Font(bold=True, color="FFFFFF")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(vertical="center")

    for r in rows:
        ws.append(r)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    widths = [16, 18, 40, 60, 8, 10, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # แผ่นสรุป
    ws2 = wb.create_sheet("Summary")
    ws2.append(["สรุปดัชนี OAS", ""])
    ws2.append(["สร้างเมื่อ", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws2.append(["ไฟล์ทั้งหมด", len(rows)])
    ws2.append([])
    ws2.append(["Section", "จำนวนไฟล์"])
    sec = {}
    for r in rows:
        sec[r[0]] = sec.get(r[0], 0) + 1
    for k in sorted(sec):
        ws2.append([k, sec[k]])
    ws2.append([])
    ws2.append(["Section / Brand", "จำนวนไฟล์"])
    sb = {}
    for r in rows:
        key = f"{r[0]} / {r[1]}"
        sb[key] = sb.get(key, 0) + 1
    for k in sorted(sb):
        ws2.append([k, sb[k]])
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 14
    for cell in (ws2["A1"], ws2["A5"], ws2["A" + str(7 + len(sec))]):
        cell.font = Font(bold=True)

    wb.save(OUT)
    print(f"OK: {len(rows)} ไฟล์ -> {OUT}")
    for k in sorted(sec):
        print(f"   {k}: {sec[k]}")


if __name__ == "__main__":
    main()
