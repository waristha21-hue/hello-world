import openpyxl
from copy import copy
from datetime import datetime

# Open the ORIGINAL template
src_path = r"C:\Users\ACER\Desktop\ขอทำใบเสนอราคา สินค้า DIAG.xlsx"
wb = openpyxl.load_workbook(src_path)
ws = wb["ใบเสนอราคา"]

# === Fill in the data ===

# B3: Hospital name
ws["B3"] = "โรงพยาบาลราชพิพัฒน์"

# F4: Date
ws["F4"] = datetime.now()

# F5: Sales rep
ws["F5"] = "วริษฐา"

# B6: Recipient
ws["B6"] = "ท่านผู้อำนวยการ"

# Clear old product rows (A9:I12)
for row in range(9, 13):
    for col in range(1, 10):
        ws.cell(row=row, column=col).value = None

# Row 9: Product - RQ9128
ws["A9"] = 1
ws["C9"] = "RIQAS Monthly Chemistry Programme\n(52 Parameters + 4 pilots, samples every month, 1 x 12 monthly cycle, 12 months subscription)"
ws["D9"] = 1
ws["E9"] = "PROGRAMME"
ws["F9"] = 22900
ws["G9"] = "=F9*1.07"
ws["H9"] = "=G9*D9"

# Row 10: Cycle info (extra row for RIQAS products)
ws["C10"] = "Cycle 24A, 24B (ม.ค.2027 - ธ.ค.2027)"

# Row 13: Product origin
ws["C13"] = "ผลิตภัณฑ์จากประเทศ : UNITED KINGDOM\nยี่ห้อ : RANDOX\nไม่มีผลิตในประเทศไทย ไม่มี มอก.\nบริษัทเป็นผู้จำหน่ายเพียงผู้เดียว"

# Summary rows - update formulas to point to row 9 only
ws["H19"] = "=H21/1.07"
ws["H20"] = "=H21-H19"
ws["H21"] = "=SUM(H9:H9)"

# Terms
ws["B25"] = "RANDOX"
ws["E25"] = "UK"

# Writer
ws["B30"] = "วริษฐา"

# Save
out_path = r"C:\Users\ACER\Desktop\ใบเสนอราคา_ราชพิพัฒน์_RIQAS_RQ9128.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
